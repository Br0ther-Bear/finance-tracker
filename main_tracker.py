import sqlite3
import openpyxl
from datetime import datetime

# Connect to SQLite database
conn = sqlite3.connect('finance_tracker.db')
cursor = conn.cursor()

# Add the new column 'type' to the existing transactions table
def add_type_column():
    cursor.execute("ALTER TABLE transactions ADD COLUMN type TEXT")
    conn.commit()

# Run this function once to add the new column to your existing table
try:
    add_type_column()
except sqlite3.OperationalError:
    # Ignore the error if the column already exists
    pass

# Create tables
cursor.execute('''CREATE TABLE IF NOT EXISTS transactions
                  (id INTEGER PRIMARY KEY, amount REAL, category TEXT, date TEXT, type TEXT)''')


def add_transaction(amount, category, date, type):
    try:
        date_obj = datetime.strptime(date, "%d-%m-%Y")
        formatted_date = date_obj.strftime("%Y-%m-%d")
        formatted_category = category.capitalize()
        
        if type == 'Income' and amount > 0:
            amount = amount  # Keep as positive
        elif type == 'Expense' and amount > 0:
            amount = -amount  # Make negative
        else:
            print("Invalid amount. Please enter a positive number.")
            return
        
        cursor.execute("INSERT INTO transactions (amount, category, date, type) VALUES (?, ?, ?, ?)",
                       (amount, formatted_category, formatted_date, type))
        conn.commit()
        print("Transaction added successfully!")
    except ValueError:
        print("Invalid date format. Please enter the date in DD-MM-YYYY format.")

def display_transactions(date_format):
    cursor.execute("SELECT * FROM transactions")
    rows = cursor.fetchall()
    print("\n")
    for row in rows:
        if date_format == 'american':
            display_date = datetime.strptime(row[3], "%Y-%m-%d").strftime("%m-%d-%Y")
        elif date_format == 'european':
            display_date = datetime.strptime(row[3], "%Y-%m-%d").strftime("%d-%m-%Y")
        else:
            display_date = row[3]
        
        # Apply color coding: Green for income, Red for expenses
        if row[4] == 'Income':
            amount_display = f"\033[92m{row[1]}\033[0m"  # Green for income
        else:
            amount_display = f"\033[91m{row[1]}\033[0m"  # Red for expenses
        
        print(f"ID: {row[0]}  |  Amount: {amount_display}  |  Category: {row[2]}  |  Date: {display_date}  |  Type: {row[4]}")

def merge_and_rename_categories(old_categories, new_category):
    for old_category in old_categories:
        cursor.execute("UPDATE transactions SET category = ? WHERE category = ?", (new_category, old_category))
    conn.commit()
    print(f"All transactions under {', '.join(old_categories)} have been merged into '{new_category}'.")

def capitalize_category_name(category_name):
    return ' '.join(word.capitalize() for word in category_name.split())

def choose_categories_to_merge():
    cursor.execute("SELECT DISTINCT category FROM transactions")
    categories = cursor.fetchall()
    categories = [category[0] for category in categories]
    selected_categories = []
    
    print("\nChoose the first category to merge:")
    for idx, category in enumerate(categories, 1):
        print(f"{idx}. {category}")
    
    choice = int(input("Enter the number of the category to merge: "))
    if 1 <= choice <= len(categories):
        selected_categories.append(categories[choice - 1])
    else:
        print("Invalid choice. Please try again.")
        return choose_categories_to_merge()
    
    while True:
        remaining_categories = [cat for cat in categories if cat not in selected_categories]
        if not remaining_categories:
            print("No more categories left to merge.")
            break

        print("\nChoose another category to merge or enter '0' to finish:")
        for idx, category in enumerate(remaining_categories, 1):
            print(f"{idx}. {category}")
        
        choice = input("Enter the number of the category to merge or '0' to finish: ")
        if choice == '0':
            break
        choice = int(choice)
        if 1 <= choice <= len(remaining_categories):
            selected_categories.append(remaining_categories[choice - 1])
        else:
            print("Invalid choice. Please try again.")
    
    return selected_categories

def prompt_for_new_category_name():
    new_category = input("Enter the new category name for the merged categories: ")
    return capitalize_category_name(new_category)

def generate_general_summary_excel():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "General Summary"

    # Add Totals Breakdown section
    sheet.append(["Totals Breakdown:"])
    totals_headers = ["Totals", "Amount"]
    sheet.append(totals_headers)

    # Fetch and add totals
    cursor.execute("SELECT SUM(amount) FROM transactions WHERE type = 'Income'")
    total_income = cursor.fetchone()[0] or 0
    cursor.execute("SELECT SUM(amount) FROM transactions WHERE type = 'Expense'")
    total_expense = abs(cursor.fetchone()[0] or 0)
    net_total = total_income - total_expense

    totals_data = [
        ["Total Income", total_income],
        ["Total Expense", total_expense],
        ["Net Total", net_total]
    ]
    for row in totals_data:
        sheet.append(row)

    # Add spacing between sections
    sheet.append([""])

    # Add Income Breakdown section
    sheet.append(["Income Breakdown:"])
    income_headers = ["Amount", "Group"]
    sheet.append(income_headers)

    # Fetch and add income data
    cursor.execute("SELECT SUM(amount), category FROM transactions WHERE type = 'Income' GROUP BY category")
    income_sums = cursor.fetchall()
    for amount, category in income_sums:
        sheet.append([amount, category])

    # Add spacing between sections
    sheet.append([""])

    # Add Expense Breakdown section
    sheet.append(["Expenses Breakdown:"])
    expense_headers = ["Amount", "Group"]
    sheet.append(expense_headers)

    # Fetch and add expense data
    cursor.execute("SELECT SUM(amount), category FROM transactions WHERE type = 'Expense' GROUP BY category")
    expense_sums = cursor.fetchall()
    for amount, category in expense_sums:
        amount_abs = abs(amount)  # Ensure the amount is positive for readability
        sheet.append([amount_abs, category])

    # Save the file
    workbook.save("general_summary.xlsx")
    print("General summary saved as 'general_summary.xlsx'.")

def generate_category_summary_excel():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Category Summary"

    # Add headers
    headers = ["Category", "Type", "Amount", "Percentage"]
    sheet.append(headers)

    # Income Breakdown
    cursor.execute("SELECT category, SUM(amount) FROM transactions WHERE type = 'Income' GROUP BY category")
    income_sums = cursor.fetchall()
    cursor.execute("SELECT SUM(amount) FROM transactions WHERE type = 'Income'")
    total_income = cursor.fetchone()[0] or 0

    for category, amount in income_sums:
        percentage = (amount / total_income) if total_income else 0
        sheet.append([category, 'Income', amount, percentage])

    # Expense Breakdown
    cursor.execute("SELECT category, SUM(amount) FROM transactions WHERE type = 'Expense' GROUP BY category")
    expense_sums = cursor.fetchall()
    cursor.execute("SELECT SUM(amount) FROM transactions WHERE type = 'Expense'")
    total_expense = cursor.fetchone()[0] or 0

    for category, amount in expense_sums:
        amount_abs = abs(amount)
        percentage = (amount_abs / total_expense) if total_expense else 0
        sheet.append([category, 'Expense', amount_abs, percentage])

    # Format percentages correctly
    for row in sheet.iter_rows(min_row=2, max_row=1+len(income_sums)+len(expense_sums), min_col=4, max_col=4):
        for cell in row:
            cell.number_format = '0.00%'

    # Save the file
    workbook.save("category_summary.xlsx")
    print("Category summary saved as 'category_summary.xlsx'.")

def generate_transactions_file():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "All Transactions"

    # Add headers for the transactions table
    headers = ["ID", "Amount", "Category", "Date", "Type"]
    sheet.append(headers)

    # Fetch all transactions
    cursor.execute("SELECT * FROM transactions")
    transactions = cursor.fetchall()

    # Add transactions data to the sheet
    for transaction in transactions:
        sheet.append(transaction)

    # Create table for all transactions
    table_ref = f"A1:E{len(transactions) + 1}"
    transactions_table = openpyxl.worksheet.table.Table(displayName="AllTransactions", ref=table_ref)
    sheet.add_table(transactions_table)

    # Save the file
    workbook.save("all_transactions.xlsx")
    print("All transactions saved as 'all_transactions.xlsx'.")


def generate_complete_summary():
    cursor.execute("SELECT SUM(amount) FROM transactions WHERE type = 'Income'")
    total_income = cursor.fetchone()[0] or 0
    
    cursor.execute("SELECT SUM(amount) FROM transactions WHERE type = 'Expense'")
    total_expense = abs(cursor.fetchone()[0] or 0)  # Convert to positive for display purposes

    net_total = total_income - total_expense
    
    print(f"\nTotal Income: {total_income}")
    print(f"Total Expense: {total_expense}")
    print(f"Net Total: {net_total}\n")
    
    # Income Breakdown
    cursor.execute("SELECT category, SUM(amount) FROM transactions WHERE type = 'Income' GROUP BY category")
    income_sums = cursor.fetchall()
    
    print("Income Breakdown:\n")
    for category, amount in income_sums:
        percentage = (amount / total_income) * 100 if total_income else 0
        print(f"Category: {category}\nAmount: {amount}\nPercentage: {percentage:.2f}%\n")
    
    # Expense Breakdown
    cursor.execute("SELECT category, SUM(amount) FROM transactions WHERE type = 'Expense' GROUP BY category")
    expense_sums = cursor.fetchall()
    
    print("Expense Breakdown:\n")
    for category, amount in expense_sums:
        amount_abs = abs(amount)  # Convert to positive for display purposes
        percentage = (amount_abs / total_expense) * 100 if total_expense else 0
        print(f"Category: {category}\nAmount: {amount_abs}\nPercentage: {percentage:.2f}%\n")

def generate_category_summary(category):
    cursor.execute("SELECT SUM(amount) FROM transactions WHERE category = ?", (category,))
    total_category_spent = cursor.fetchone()[0] or 0
    print(f"\nTotal Amount Spent in {category}: {total_category_spent}")

def merge_categories_menu():
    old_categories = choose_categories_to_merge()
    new_category = prompt_for_new_category_name()
    merge_and_rename_categories(old_categories, new_category)


def delete_transaction(transaction_id):
    cursor.execute("DELETE FROM transactions WHERE id = ?", (transaction_id,))
    conn.commit()
    if cursor.rowcount == 0:
        print(f"No transaction found with ID {transaction_id}.")
    else:
        print("Transaction deleted successfully!")

def delete_transactions_by_date_range(start_date, end_date):
    try:
        start_date_obj = datetime.strptime(start_date, "%d-%m-%Y").strftime("%Y-%m-%d")
        end_date_obj = datetime.strptime(end_date, "%d-%m-%Y").strftime("%Y-%m-%d")
        cursor.execute("DELETE FROM transactions WHERE date BETWEEN ? AND ?", (start_date_obj, end_date_obj))
        conn.commit()
        print(f"Transactions from {start_date} to {end_date} have been deleted.")
    except ValueError:
        print("Invalid date format. Please enter dates in DD-MM-YYYY format.")

def delete_all_transactions():
    cursor.execute("DELETE FROM transactions")
    conn.commit()
    print("All transactions have been deleted.")

def delete_transactions_by_category(category):
    cursor.execute("DELETE FROM transactions WHERE category = ?", (category,))
    conn.commit()
    print(f"All transactions in the category '{category}' have been deleted.")

def choose_date_format():
    while True:
        print("\nChoose Date Format:")
        print("1. American (MM-DD-YYYY)")
        print("2. European (DD-MM-YYYY)")
        print("3. Return to Main Menu")
        choice = input("Enter your choice: ")
        
        if choice == '1':
            return 'american'
        elif choice == '2':
            return 'european'
        elif choice == '3':
            return 'main_menu'
        else:
            print("Invalid choice. Please try again.")

def choose_summary_type():
    while True:
        print("\nChoose Summary Type:")
        print("1. Complete Summary")
        print("2. Summary by Category")
        print("3. Return to Main Menu")
        choice = input("Enter your choice: ")
        
        if choice == '1':
            return 'complete'
        elif choice == '2':
            return 'category'
        elif choice == '3':
            return 'main_menu'
        else:
            print("Invalid choice. Please try again.")

def choose_category():
    cursor.execute("SELECT DISTINCT category FROM transactions")
    categories = cursor.fetchall()
    categories = [category[0] for category in categories]
    while True:
        print("\nChoose a Category:")
        for idx, category in enumerate(categories, 1):
            print(f"{idx}. {category}")
        choice = int(input("Enter your choice: "))
        
        if 1 <= choice <= len(categories):
            return categories[choice - 1]
        else:
            print("Invalid choice. Please try again.")

def summary_menu():
    while True:
        print("\nSummary Menu:")
        print("1. General Summary")
        print("2. Summary by Group")
        print("3. Return to Main Menu")
        choice = input("Enter your choice: ")

        if choice == '1':
            summary_submenu("general")
        elif choice == '2':
            summary_submenu("category")
        elif choice == '3':
            break
        else:
            print("Invalid choice. Please try again.")

def summary_submenu(summary_type):
    while True:
        print("\nSummary Options:")
        print("1. View Summary")
        print("2. Generate Summary File")
        print("3. Return to Summary Menu")
        choice = input("Enter your choice: ")

        if choice == '1':
            if summary_type == "general":
                generate_complete_summary()
            elif summary_type == "category":
                generate_category_summary()
        elif choice == '2':
            if summary_type == "general":
                generate_general_summary_excel()
            elif summary_type == "category":
                generate_category_summary_excel()
        elif choice == '3':
            break
        else:
            print("Invalid choice. Please try again.")

def add_transaction_menu():
    while True:
        print("\nAdd Transaction Menu:")
        print("1. Add Income")
        print("2. Add Expense")
        print("3. Return to Transaction Menu")
        choice = input("Enter your choice: ")

        if choice == '1':
            type = 'Income'
            amount = float(input("Enter amount: "))
            category = input("Enter category: ")
            date = input("Enter date (DD-MM-YYYY): ")
            add_transaction(amount, category, date, type)
        elif choice == '2':
            type = 'Expense'
            amount = float(input("Enter amount: "))
            category = input("Enter category: ")
            date = input("Enter date (DD-MM-YYYY): ")
            add_transaction(amount, category, date, type)
        elif choice == '3':
            break
        else:
            print("Invalid choice. Please try again.")

def transaction_menu():
    while True:
        print("\nTransaction Menu:")
        print("1. Add a Transaction")
        print("2. Delete a Transaction")
        print("3. Merge Transaction Groups")
        print("4. Return to Main Menu")
        choice = input("Enter your choice: ")

        if choice == '1':
            add_transaction_menu()
        elif choice == '2':
            delete_transaction_menu()
        elif choice == '3':
            merge_categories_menu()
        elif choice == '4':
            break
        else:
            print("Invalid choice. Please try again.")


def delete_transaction_menu():
    while True:
        print("\nDelete Transaction Menu:")
        print("1. Delete by ID")
        print("2. Delete by Date Range")
        print("3. Delete by Category")
        print("4. Delete All Transactions")
        print("5. Return to Transaction Menu")
        choice = input("Enter your choice: ")

        if choice == '1':
            transaction_id = int(input("Enter transaction ID to delete: "))
            delete_transaction(transaction_id)
        elif choice == '2':
            start_date = input("Enter start date (DD-MM-YYYY): ")
            end_date = input("Enter end date (DD-MM-YYYY): ")
            delete_transactions_by_date_range(start_date, end_date)
        elif choice == '3':
            category = choose_category()
            delete_transactions_by_category(category)
        elif choice == '4':
            confirm = input("Are you sure you want to delete all transactions? (yes/no): ").strip().lower()
            if confirm == 'yes':
                delete_all_transactions()
            else:
                print("Deletion cancelled.")
        elif choice == '5':
            break
        else:
            print("Invalid choice. Please try again.")

def main_menu():
    while True:
        print("\nFinancial Tracker Menu:")
        print("1. Transactions")
        print("2. View Transactions")
        print("3. Generate Summary")
        print("4. Merge Categories")
        print("5. Generate Transactions File")  # New option
        print("6. Exit")
        choice = input("Enter your choice: ")

        if choice == '1':
            transaction_menu()
        elif choice == '2':
            date_format = choose_date_format()
            if date_format == 'main_menu':
                continue
            display_transactions(date_format)
        elif choice == '3':
            summary_menu()
        elif choice == '4':
            merge_categories_menu()
        elif choice == '5':
            generate_transactions_file()  # Call the new function
        elif choice == '6':
            break
        else:
            print("Invalid choice. Please try again.")


# Running the main menu
main_menu()

# Closing the connection
conn.close()
